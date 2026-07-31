#!/usr/bin/env python3
"""
excel_smart.py — 复杂 Excel → Markdown（L1 清洗 + L2 结构感知）

L2（openpyxl，.xlsx/.xlsm）：
  - 合并单元格向下/向右填充
  - 空行切分为多张表
  - 双行表头合并为「上级 / 下级」
  - 删除全空列、清理 NaN/Unnamed 噪声

L1（.xls 或 openpyxl 失败回退）：由调用方走 markitdown + clean_markdown_tables。

Part of huashu-md-html skill.
"""
from __future__ import annotations

import re
from datetime import date, datetime, time
from decimal import Decimal
from pathlib import Path
from typing import Any


_EMPTY_CELL_TOKENS = frozenset({"", "nan", "none", "nat", "<na>", "null"})
_UNNAMED_RE = re.compile(r"^Unnamed:\s*\d+$", re.IGNORECASE)
_OPENXML_SUFFIXES = frozenset({".xlsx", ".xlsm"})


def supports_openpyxl(path: Path) -> bool:
    return path.suffix.lower() in _OPENXML_SUFFIXES


def cell_to_text(value: Any) -> str:
    """把单元格值规范成适合 Markdown 的纯文本。"""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, float):
        if value != value:  # NaN
            return ""
        if value == int(value) and abs(value) < 1e15:
            return str(int(value))
        return str(value)
    if isinstance(value, Decimal):
        s = format(value, "f")
        if "." in s:
            s = s.rstrip("0").rstrip(".")
        return s
    if isinstance(value, datetime):
        if value.hour or value.minute or value.second or value.microsecond:
            return value.strftime("%Y-%m-%d %H:%M:%S")
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.strftime("%H:%M:%S")
    s = str(value).strip()
    if s.lower() in _EMPTY_CELL_TOKENS:
        return ""
    if _UNNAMED_RE.fullmatch(s):
        return ""
    # Markdown 表格内竖线转义
    return s.replace("|", "\\|").replace("\n", " ").replace("\r", "")


def _row_all_empty(row: list[str]) -> bool:
    return all(not c.strip() for c in row)


def _trim_grid(grid: list[list[str]]) -> list[list[str]]:
    """去掉四周全空边框。"""
    if not grid:
        return grid
    # 去尾部空行
    while grid and _row_all_empty(grid[-1]):
        grid.pop()
    # 去头部空行
    while grid and _row_all_empty(grid[0]):
        grid.pop(0)
    if not grid:
        return grid
    width = max(len(r) for r in grid)
    for r in grid:
        while len(r) < width:
            r.append("")
        del r[width:]
    # 去左侧全空列
    while width and all(not r[0].strip() for r in grid):
        for r in grid:
            r.pop(0)
        width -= 1
    # 去右侧全空列
    while width and all(not r[-1].strip() for r in grid):
        for r in grid:
            r.pop()
        width -= 1
    return grid


def _drop_empty_columns(rows: list[list[str]]) -> list[list[str]]:
    if not rows:
        return rows
    width = max(len(r) for r in rows)
    for r in rows:
        while len(r) < width:
            r.append("")
    keep = [j for j in range(width) if any((r[j] if j < len(r) else "").strip() for r in rows)]
    if not keep:
        return rows
    return [[r[j] for j in keep] for r in rows]


def _looks_like_data_row(row: list[str]) -> bool:
    """首个非空单元格像数字 → 更像数据行。"""
    for c in row:
        if c.strip():
            return _numericish(c)
    return False


def _split_blocks(grid: list[list[str]]) -> list[list[list[str]]]:
    """
    空行分表：
    - 连续 ≥2 空行：一定切开
    - 单空行：仅当下一行像新表头、且当前块末行不像表头时切开
    """
    blocks: list[list[list[str]]] = []
    current: list[list[str]] = []
    blank_streak = 0
    for row in grid:
        if _row_all_empty(row):
            blank_streak += 1
            continue
        should_split = False
        if current and blank_streak >= 1:
            if blank_streak >= 2:
                should_split = True
            elif (
                len(current) >= 2
                and _looks_like_header_row(row)
                and (
                    _looks_like_data_row(current[-1])
                    or not _looks_like_header_row(current[-1])
                )
            ):
                should_split = True
        if should_split:
            blocks.append(current)
            current = []
        blank_streak = 0
        current.append(list(row))
    if current:
        blocks.append(current)
    return blocks


def _carry_headers(blocks: list[list[list[str]]]) -> list[list[list[str]]]:
    """无表头的数据块继承上一块的表头；并对双行表头做合并。"""
    out: list[list[list[str]]] = []
    last_header: list[str] | None = None
    for block in blocks:
        if not block:
            continue
        processed = _merge_double_header([list(r) for r in block])
        if _looks_like_header_row(processed[0]):
            last_header = list(processed[0])
            out.append(processed)
        elif last_header is not None and _looks_like_data_row(processed[0]):
            out.append([list(last_header)] + processed)
        else:
            out.append(processed)
    return out


def _numericish(text: str) -> bool:
    s = text.strip().replace(",", "").replace("%", "")
    if not s:
        return False
    try:
        float(s)
        return True
    except ValueError:
        return False


def _looks_like_header_row(row: list[str]) -> bool:
    """启发式：非空单元格偏文本、且不全是数字。"""
    cells = [c.strip() for c in row if c.strip()]
    if not cells:
        return False
    numeric = sum(1 for c in cells if _numericish(c))
    return numeric / len(cells) <= 0.4


def _merge_double_header(block: list[list[str]]) -> list[list[str]]:
    """若前两行都像表头，合并为单行「上级 / 下级」。"""
    if len(block) < 3:
        return block
    r0, r1 = block[0], block[1]
    if not (_looks_like_header_row(r0) and _looks_like_header_row(r1)):
        return block
    # 第二行若几乎全空，不合并（可能是标题下的空行残留）
    if sum(1 for c in r1 if c.strip()) < max(1, sum(1 for c in r0 if c.strip()) // 3):
        return block
    width = max(len(r0), len(r1))
    merged: list[str] = []
    for j in range(width):
        a = (r0[j] if j < len(r0) else "").strip()
        b = (r1[j] if j < len(r1) else "").strip()
        if a and b and a != b:
            merged.append(f"{a} / {b}")
        else:
            merged.append(a or b)
    return [merged] + block[2:]


def _to_markdown_table(block: list[list[str]]) -> str:
    block = _drop_empty_columns(block)
    block = _trim_grid(block)
    if not block:
        return ""

    width = max(len(r) for r in block)
    for r in block:
        while len(r) < width:
            r.append("")

    # 单行：当作只有表头的表
    header = block[0]
    body = block[1:] if len(block) > 1 else []
    # 若首行不像表头而后续有数据，造 Col1..N
    if body and not _looks_like_header_row(header):
        body = block
        header = [f"Col{j + 1}" for j in range(width)]

    lines = [
        "| " + " | ".join(header) + " |",
        "| " + " | ".join("---" for _ in header) + " |",
    ]
    for row in body:
        lines.append("| " + " | ".join(row) + " |")
    return "\n".join(lines)


def _load_sheet_grid(ws) -> list[list[str]]:
    """读取 worksheet 为二维文本网格，并展开合并单元格。"""
    if ws.max_row is None or ws.max_column is None:
        return []
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    if max_row < 1 or max_col < 1:
        return []

    # 限制极端宽表（防内存爆炸）；通常业务表远小于此
    max_row = min(max_row, 50000)
    max_col = min(max_col, 200)

    grid: list[list[str]] = [
        [cell_to_text(ws.cell(row=r, column=c).value) for c in range(1, max_col + 1)]
        for r in range(1, max_row + 1)
    ]

    for merged in ws.merged_cells.ranges:
        min_r, min_c, max_r, max_c = merged.min_row, merged.min_col, merged.max_row, merged.max_col
        if min_r > max_row or min_c > max_col:
            continue
        max_r = min(max_r, max_row)
        max_c = min(max_c, max_col)
        fill = grid[min_r - 1][min_c - 1]
        for r in range(min_r, max_r + 1):
            for c in range(min_c, max_c + 1):
                if r == min_r and c == min_c:
                    continue
                # 仅填充原本为空的格子，避免覆盖误判
                if not grid[r - 1][c - 1].strip():
                    grid[r - 1][c - 1] = fill

    return _trim_grid(grid)


def convert_workbook_to_md(path: Path) -> str:
    """
    将本地 .xlsx/.xlsm 转为 Markdown。
    需要 openpyxl；失败时抛出异常由调用方回退。
    """
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError(
            "excel-smart 需要 openpyxl：python -m pip install openpyxl",
        ) from exc

    # data_only=True 取公式缓存值；无缓存时为 None（比公式字符串更适合喂模型）
    wb = load_workbook(path, data_only=True, read_only=False)
    parts: list[str] = []
    try:
        for name in wb.sheetnames:
            ws = wb[name]
            if getattr(ws, "sheet_state", "visible") == "hidden":
                continue
            grid = _load_sheet_grid(ws)
            if not grid:
                continue
            blocks = _carry_headers(_split_blocks(grid))
            if not blocks:
                continue
            # 先渲染，避免空块导致 ### Table 编号空洞/误标
            tables = [md for md in (_to_markdown_table(b) for b in blocks) if md]
            if not tables:
                continue
            parts.append(f"## {name}")
            for idx, table_md in enumerate(tables, start=1):
                if len(tables) > 1:
                    parts.append(f"### Table {idx}")
                parts.append(table_md)
                parts.append("")
    finally:
        wb.close()
    if not parts:
        return f"# {path.name}\n\n（工作簿无可见数据）\n"
    return f"# {path.stem}\n\n" + "\n".join(parts).rstrip() + "\n"


def try_convert(path: Path) -> str | None:
    """成功返回 md；不支持或失败返回 None。"""
    if not supports_openpyxl(path):
        return None
    try:
        return convert_workbook_to_md(path)
    except Exception:  # noqa: BLE001 — 调用方回退 markitdown
        return None
